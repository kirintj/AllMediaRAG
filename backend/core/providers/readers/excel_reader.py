"""Excel/CSV 文件读取器"""
from __future__ import annotations
import os
import csv
import io


class ExcelReader:
    """读取 .xlsx 和 .csv 文件，转为 Markdown 表格格式"""

    def read(self, file_path: str) -> str:
        ext = os.path.splitext(file_path)[1].lower()
        if ext == ".csv":
            return self._read_csv(file_path)
        return self._read_excel(file_path)

    def supported_extensions(self) -> list[str]:
        return [".xlsx", ".csv"]

    def _read_excel(self, path: str) -> str:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        parts = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            # 过滤全空行
            rows = [r for r in rows if any(c is not None for c in r)]
            if not rows:
                continue
            header = " | ".join(str(c) if c is not None else "" for c in rows[0])
            parts.append(f"## {sheet}\n| {header} |")
            parts.append("|" + " --- |" * len(rows[0]))
            for row in rows[1:]:
                cells = [str(c) if c is not None else "" for c in row]
                # 补齐列数
                while len(cells) < len(rows[0]):
                    cells.append("")
                parts.append("| " + " | ".join(cells) + " |")
        wb.close()
        return "\n".join(parts)

    def _read_csv(self, path: str) -> str:
        # 自动检测编码
        for encoding in ["utf-8", "utf-8-sig", "gbk", "latin-1"]:
            try:
                with open(path, "r", encoding=encoding) as f:
                    content = f.read()
                break
            except (UnicodeDecodeError, UnicodeError):
                continue
        else:
            with open(path, "r", encoding="utf-8", errors="replace") as f:
                content = f.read()

        reader = csv.reader(io.StringIO(content))
        rows = list(reader)
        if not rows:
            return ""

        header = " | ".join(rows[0])
        parts = [f"| {header} |", "|" + " --- |" * len(rows[0])]
        for row in rows[1:]:
            while len(row) < len(rows[0]):
                row.append("")
            parts.append("| " + " | ".join(row) + " |")
        return "\n".join(parts)
