"""标签知识库管理

解析用户上传的标签文件（Excel/CSV），存入 ES 作为标签 chunk。
与 RAGFlow 的 tag.py 对齐。
"""
from __future__ import annotations

import os
import uuid
import logging
import csv
import io
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from core.providers.elasticsearch_store import ElasticsearchStore

logger = logging.getLogger(__name__)


class TagKBManager:
    """标签知识库管理器"""

    def __init__(self, doc_store: ElasticsearchStore):
        self._store = doc_store

    def parse_tag_file(self, file_path: str) -> list[dict]:
        """解析标签文件（Excel/CSV），返回 [{content, tags: [str]}]"""
        ext = os.path.splitext(file_path)[1].lower()
        if ext == ".csv":
            return self._parse_csv(file_path)
        elif ext in (".xlsx", ".xls"):
            return self._parse_excel(file_path)
        else:
            raise ValueError(f"Unsupported tag file format: {ext}")

    def ingest_tag_file(self, file_path: str, tag_kb_id: str) -> int:
        """解析标签文件并存入 ES

        Args:
            file_path: 标签文件路径
            tag_kb_id: 标签知识库 ID

        Returns:
            存入的标签 chunk 数量
        """
        entries = self.parse_tag_file(file_path)
        if not entries:
            return 0

        source = os.path.basename(file_path)
        rows = []
        for entry in entries:
            tags = [t.strip().replace(".", "_") for t in entry["tags"] if t.strip()]
            if not tags:
                continue
            rows.append({
                "id": str(uuid.uuid4()),
                "text": self._tokenize(entry["content"]),
                "text_raw": entry["content"],
                "embedding": [],  # 标签 chunk 不需要 embedding
                "source": source,
                "chunk_type": "tag",
                "tag_kwd": tags,
                "metadata": {"tag_kb_id": tag_kb_id},
            })

        if rows:
            self._store.insert(rows)
            logger.info("Ingested %d tag chunks for KB %s", len(rows), tag_kb_id)

        return len(rows)

    def get_all_tags(self, tag_kb_id: str) -> dict[str, float]:
        """从 ES 聚合获取标签集（带平滑频率）

        Returns:
            {tag: smoothed_frequency}
        """
        # ES aggregation on tag_kwd field
        try:
            result = self._store._client.search(
                index=self._store._index_name,
                body={
                    "size": 0,
                    "query": {
                        "bool": {
                            "filter": [
                                {"term": {"chunk_type": "tag"}},
                                {"term": {"metadata.tag_kb_id": tag_kb_id}},
                            ]
                        }
                    },
                    "aggs": {
                        "tags": {
                            "terms": {"field": "tag_kwd", "size": 1000}
                        }
                    }
                }
            )
            buckets = result["aggregations"]["tags"]["buckets"]
            total = sum(b["doc_count"] for b in buckets)
            S = 1000  # smoothing constant (same as RAGFlow)
            return {b["key"]: (b["doc_count"] + 1) / (total + S) for b in buckets}
        except Exception as e:
            logger.warning("Failed to aggregate tags for KB %s: %s", tag_kb_id, e)
            return {}

    def delete_tag_kb(self, tag_kb_id: str):
        """删除标签知识库"""
        self._store.delete({"chunk_type": "tag", "metadata.tag_kb_id": tag_kb_id})

    def list_tag_kbs(self) -> list[dict]:
        """列出所有标签知识库"""
        try:
            result = self._store._client.search(
                index=self._store._index_name,
                body={
                    "size": 0,
                    "query": {"term": {"chunk_type": "tag"}},
                    "aggs": {
                        "by_kb": {
                            "terms": {"field": "metadata.tag_kb_id", "size": 100}
                        }
                    }
                }
            )
            kbs = []
            for bucket in result["aggregations"]["by_kb"]["buckets"]:
                kbs.append({"id": bucket["key"], "chunk_count": bucket["doc_count"]})
            return kbs
        except Exception as e:
            logger.warning("Failed to list tag KBs: %s", e)
            return []

    def _parse_csv(self, path: str) -> list[dict]:
        entries = []
        for encoding in ["utf-8", "utf-8-sig", "gbk", "latin-1"]:
            try:
                with open(path, "r", encoding=encoding) as f:
                    reader = csv.DictReader(f)
                    for row in reader:
                        content = row.get("content", "").strip()
                        tags_str = row.get("tags", "")
                        tags = [t.strip() for t in tags_str.split(",") if t.strip()]
                        if content and tags:
                            entries.append({"content": content, "tags": tags})
                return entries
            except (UnicodeDecodeError, UnicodeError):
                continue
        return entries

    def _parse_excel(self, path: str) -> list[dict]:
        from openpyxl import load_workbook
        entries = []
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if len(rows) < 2:
            return entries
        header = [str(c).strip().lower() if c else "" for c in rows[0]]
        content_idx = header.index("content") if "content" in header else 0
        tags_idx = header.index("tags") if "tags" in header else 1
        for row in rows[1:]:
            content = str(row[content_idx]).strip() if row[content_idx] else ""
            tags_str = str(row[tags_idx]) if row[tags_idx] else ""
            tags = [t.strip() for t in tags_str.split(",") if t.strip()]
            if content and tags:
                entries.append({"content": content, "tags": tags})
        return entries

    def _tokenize(self, text: str) -> str:
        import jieba
        return " ".join(jieba.cut(text))
